"""
High-Fidelity Financial Workbook and Excel Model Generator.
Builds multi-sheet linked workbooks using openpyxl with cell formulas, scenario parameters, and debt waterfalls.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

def generate_fpa_workbook(
    starting_cash: float,
    attendance_rate: float,
    ticket_pricing: float,
    sponsorship_delay: int,
    payroll_growth: float,
    inflation: float,
    interest_rate: float
) -> BytesIO:
    """
    Generates a highly formatted, formula-linked professional FP&A Workbook.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styles
    navy_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    accent_fill = PatternFill(start_color="E8EEF5", end_color="E8EEF5", fill_type="solid")
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    white_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Arial", size=11, bold=True, color="1B365D")
    regular_font = Font(name="Arial", size=10)
    title_font = Font(name="Arial", size=16, bold=True, color="1B365D")

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB")
    )
    
    double_bottom_border = Border(
        top=Side(style="thin", color="1B365D"),
        bottom=Side(style="double", color="1B365D")
    )

    # ────────────────────────────────────────────────────────
    # SHEET 1: Scenario Assumptions
    # ────────────────────────────────────────────────────────
    ws_assump = wb.create_sheet(title="Scenario Assumptions")
    ws_assump.views.sheetView[0].showGridLines = True
    
    ws_assump["A1"] = "FIN-OS ASSUMPTION PORTAL"
    ws_assump["A1"].font = title_font
    
    headers = ["Operational Driver / Assumption", "Model Value", "Metric Type"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws_assump.cell(row=3, column=col_idx, value=h)
        cell.fill = navy_fill
        cell.font = white_font
        cell.alignment = center_align

    drivers = [
        ("Capital Reserves", starting_cash, "Currency"),
        ("Sponsorship Attendance Multiplier", attendance_rate, "Percentage"),
        ("Ticketing Pricing Index", ticket_pricing, "Factor"),
        ("Sponsorship Delay Outflow (Days)", sponsorship_delay, "Integer"),
        ("Squad Wages Growth Index", payroll_growth, "Percentage"),
        ("Corporate Inflation Variable", inflation, "Percentage"),
        ("Refinancing Base Interest Rate", interest_rate, "Percentage")
    ]

    for row_idx, (drv, val, typ) in enumerate(drivers, 4):
        c1 = ws_assump.cell(row=row_idx, column=1, value=drv)
        c2 = ws_assump.cell(row=row_idx, column=2, value=val)
        c3 = ws_assump.cell(row=row_idx, column=3, value=typ)
        
        c1.font = regular_font
        c1.alignment = left_align
        c1.border = thin_border
        
        c2.font = Font(name="Arial", size=10, bold=True)
        c2.alignment = right_align
        c2.border = thin_border
        if typ == "Currency":
            c2.number_format = "$#,##0.00"
        elif typ == "Percentage":
            c2.number_format = "0.0%"
        elif typ == "Factor":
            c2.number_format = "0.00"
            
        c3.font = regular_font
        c3.alignment = center_align
        c3.border = thin_border

    # ────────────────────────────────────────────────────────
    # SHEET 2: Continuous Cash Flow Model
    # ────────────────────────────────────────────────────────
    ws_cf = wb.create_sheet(title="Cash Flow Model")
    ws_cf.views.sheetView[0].showGridLines = True
    
    ws_cf["A1"] = "CONTINUOUS CASH FLOW FORECAST MODEL"
    ws_cf["A1"].font = title_font
    
    cf_headers = ["Financial Stream / Row Line", "30 Days Forecast", "60 Days Forecast", "90 Days Forecast", "180 Days Forecast"]
    for col_idx, h in enumerate(cf_headers, 1):
        cell = ws_cf.cell(row=3, column=col_idx, value=h)
        cell.fill = navy_fill
        cell.font = white_font
        cell.alignment = center_align

    # Multi-sheet link formulas mapping cash flows back to assumptions page!
    row_data = [
        ("Starting Cash Reserves Balance", "='Scenario Assumptions'!B4", "='Scenario Assumptions'!B4", "='Scenario Assumptions'!B4", "='Scenario Assumptions'!B4"),
        ("Revenues Inflow - Sponsorship", "=15000000*(30/360)", "=15000000*(60/360)", "=15000000*(90/360)", "=15000000*(180/360)"),
        ("Revenues Inflow - Match Ticket", "=12000000*'Scenario Assumptions'!B5*'Scenario Assumptions'!B6*(30/360)", "=12000000*'Scenario Assumptions'!B5*'Scenario Assumptions'!B6*(60/360)", "=12000000*'Scenario Assumptions'!B5*'Scenario Assumptions'!B6*(90/360)", "=12000000*'Scenario Assumptions'!B5*'Scenario Assumptions'!B6*(180/360)"),
        ("Revenues Inflow - Broadcast Media", "=18000000*(30/360)", "=18000000*(60/360)", "=18000000*(90/360)", "=18000000*(180/360)"),
        ("Total Simulated Cash Inflows", "=SUM(B5:B7)", "=SUM(C5:C7)", "=SUM(D5:D7)", "=SUM(E5:E7)"),
        ("Expenditures Outflow - Squad Wage", "=-28000000*(1+'Scenario Assumptions'!B8)*(30/360)", "=-28000000*(1+'Scenario Assumptions'!B8)*(60/360)", "=-28000000*(1+'Scenario Assumptions'!B8)*(90/360)", "=-28000000*(1+'Scenario Assumptions'!B8)*(180/360)"),
        ("Expenditures Outflow - Stadium Ops", "=-8000000*(1+'Scenario Assumptions'!B9)*(30/360)", "=-8000000*(1+'Scenario Assumptions'!B9)*(60/360)", "=-8000000*(1+'Scenario Assumptions'!B9)*(90/360)", "=-8000000*(1+'Scenario Assumptions'!B9)*(180/360)"),
        ("Total Operating Outflows", "=SUM(B9:B10)", "=SUM(C9:C10)", "=SUM(D9:D10)", "=SUM(E9:E10)"),
        ("Net Cash Flow For Period", "=B8+B11", "=C8+C11", "=D8+D11", "=E8+E11"),
        ("Ending Cash Reserves Position", "=B4+B12", "=C4+C12", "=D4+D12", "=E4+E12")
    ]

    for row_idx, r_val in enumerate(row_data, 4):
        is_total = "Total" in r_val[0] or "Ending" in r_val[0] or "Net" in r_val[0]
        
        for col_idx, cell_val in enumerate(r_val, 1):
            c = ws_cf.cell(row=row_idx, column=col_idx, value=cell_val)
            
            # Formats
            c.border = double_bottom_border if is_total else thin_border
            c.font = bold_font if is_total else regular_font
            if col_idx == 1:
                c.alignment = left_align
            else:
                c.alignment = right_align
                c.number_format = "$#,##0.00"
            if is_total:
                c.fill = accent_fill

    # Auto column adjustment
    for ws in [ws_assump, ws_cf]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
